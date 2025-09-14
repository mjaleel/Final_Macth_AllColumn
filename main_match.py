# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from rapidfuzz import fuzz
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

# تعيين عنوان الصفحة والخط
st.set_page_config(page_title="أداة مطابقة البيانات", layout="centered")
st.markdown("""
    <style>
        .stButton button {
            background-color: #4CAF50;
            color: white;
            padding: 10px 24px;
            font-size: 16px;
            border-radius: 8px;
        }
        .stFileUploader > label {
            font-size: 18px;
            font-weight: bold;
        }
        .st-emotion-cache-1c99sb8 {
            text-align: right;
            direction: rtl;
        }
    </style>
""", unsafe_allow_html=True)

# ----------------- وظائف معالجة البيانات -----------------

def normalize_name(name):
    """
    يقوم بتطبيع الاسم لإزالة المسافات الزائدة، الأحرف الخاصة، والتشكيل.
    """
    if pd.isnull(name):
        return ""
    name = str(name).strip().replace("ه", "ة").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    name = re.sub(r'(عبد)([^\s])', r'\1 \2', name)
    name = re.sub(r'[\u064b-\u0652]', '', name)
    return " ".join(name.split()).lower()

def is_first_three_words_match(name1, name2):
    """
    يتحقق مما إذا كانت أول ثلاث كلمات متطابقة.
    """
    words1 = normalize_name(name1).split()
    words2 = normalize_name(name2).split()
    length = min(len(words1), len(words2), 3)
    return all(words1[i] == words2[i] for i in range(length))

def comprehensive_match(names_file, database_file, selected_columns):
    """
    يقوم بعملية المطابقة الشاملة.
    """
    names_df = pd.read_excel(names_file)
    database_df = pd.read_excel(database_file)
    
    names_df["normalized_name"] = names_df["اسم الموظف"].apply(normalize_name)
    database_df["normalized_name"] = database_df["اسم الموظف"].apply(normalize_name)
    database_df = database_df.drop_duplicates(subset=["normalized_name"])
    
    database_map = database_df.set_index("normalized_name")[selected_columns].to_dict(orient="index")
    
    matched_results = []
    for original_name, normalized_name in zip(names_df["اسم الموظف"], names_df["normalized_name"]):
        best_match = None
        best_score = 0
        
        # المرحلة 1: البحث عن أفضل تطابق باستخدام fuzzy matching
        for db_name in database_map.keys():
            score = fuzz.ratio(normalized_name, db_name)
            if score > best_score:
                best_score = score
                best_match = db_name

        match_data = None
        if best_score >= 85 and (is_first_three_words_match(normalized_name, best_match) or best_match.startswith(normalized_name)):
            match_data = database_map[best_match]
        else:
            # المرحلة 2: إذا فشل التطابق في المرحلة الأولى، البحث عن تطابق يبدأ بنفس الاسم
            for db_name in database_map.keys():
                if db_name.startswith(normalized_name):
                    match_data = database_map[db_name]
                    best_match = db_name
                    best_score = fuzz.ratio(normalized_name, best_match)
                    break

        result_row = {
            "الاسم الأصلي": original_name,
            "الاسم المطابق": match_data["اسم الموظف"] if match_data and "اسم الموظف" in match_data else "",
            "نسبة التطابق": f"{round(best_score)}%" if match_data else "", # السطر الذي تم تعديله
            "ملاحظة": "✅ تطابق دقيق" if match_data else "❌ لم يتم العثور على تطابق"
        }
        
        for col in selected_columns:
            if col != "اسم الموظف":
                result_row[col] = match_data.get(col, "") if match_data else ""
        
        matched_results.append(result_row)
    
    results_df = pd.DataFrame(matched_results)
    
    # التحقق من تكرار الـ IBAN
    if "Iban" in selected_columns:
        iban_counts = results_df["Iban"].value_counts()
        results_df["تنبيه"] = results_df["Iban"].apply(
            lambda x: "⚠️ مكرر" if pd.notnull(x) and x in iban_counts and iban_counts[x] > 1 else ""
        )
    
    return results_df

def apply_excel_formatting(df):
    """
    يطبق التنسيق الشرطي ويعدل عرض الأعمدة في ملف Excel.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='النتائج')
    
    wb = load_workbook(output)
    ws = wb.active
    
    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
        
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    note_col_idx = df.columns.get_loc("ملاحظة") + 1
    
    for row_idx, row_data in enumerate(df.itertuples(), 2):
        note_val = getattr(row_data, 'ملاحظة', None)
        
        if note_val and "❌" in note_val:
            for cell in ws[row_idx]:
                cell.fill = red_fill
        
        if "تنبيه" in df.columns:
            alert_val = getattr(row_data, 'تنبيه', None)
            if alert_val and "⚠️" in alert_val:
                for cell in ws[row_idx]:
                    cell.fill = yellow_fill

    output.seek(0)
    wb.save(output)
    output.seek(0)
    return output

# ----------------- واجهة المستخدم Streamlit -----------------

st.title("أداة مطابقة بيانات Excel")
st.markdown("يرجى رفع ملف الأسماء وملف قاعدة البيانات لبدء عملية المطابقة الشاملة.")

names_file = st.file_uploader("اختر ملف الأسماء (يحتوي على عمود 'اسم الموظف')", type=["xlsx"])
database_file = st.file_uploader("اختر ملف قاعدة البيانات (يحتوي على عمود 'اسم الموظف')", type=["xlsx"])

if names_file and database_file:
    try:
        # قراءة الأعمدة من ملف قاعدة البيانات لاختيارها
        database_df_temp = pd.read_excel(database_file)
        all_columns = database_df_temp.columns.tolist()
        
        # إزالة الأعمدة التي لا نحتاجها في الاختيار
        if "اسم الموظف" in all_columns:
            all_columns.remove("اسم الموظف")
        
        selected_columns_user = st.multiselect(
            "اختر الأعمدة الإضافية التي تريد إضافتها إلى النتائج:",
            all_columns
        )
        
        # التأكد من أن "اسم الموظف" دائماً موجود
        selected_columns = selected_columns_user + ["اسم الموظف"]
        
        if st.button("بدء المطابقة الشاملة"):
            with st.spinner("جاري المعالجة... قد يستغرق الأمر بعض الوقت."):
                results_df = comprehensive_match(names_file, database_file, selected_columns)
                formatted_excel = apply_excel_formatting(results_df)

                st.success("تمت عملية المطابقة بنجاح!")
                st.download_button(
                    label="تنزيل ملف النتائج",
                    data=formatted_excel,
                    file_name="نتائج_المطابقة.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    except Exception as e:
        st.error(f"حدث خطأ: {e}")
        st.info("يرجى التأكد من أن ملفات Excel تحتوي على عمود باسم 'اسم الموظف' وأن التنسيق صحيح.")
